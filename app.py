import streamlit as st
import requests
import re
import time
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── Configuration ────────────────────────────────────────────────────────────

GALLICA_SRU    = "https://gallica.bnf.fr/SRU"
GALLICA_BASE   = "https://gallica.bnf.fr"
TIMEOUT        = 15
MAX_ISSUES     = 500

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; GallicaHarvester/1.0; "
        "+https://github.com/user/gallica-harvester)"
    )
}

# ─── Utilitaires ──────────────────────────────────────────────────────────────

def extract_ark(url: str) -> str | None:
    match = re.search(r"ark:/12148/([a-z0-9]+)", url)
    if match:
        return f"ark:/12148/{match.group(1)}"
    return None


def build_gallica_url(ark_value: str) -> str:
    ark_value = ark_value.strip()
    if ark_value.startswith("http") and "ark:/12148/" in ark_value:
        return re.sub(r"\.r=[^\s/]*", "", ark_value)
    if ark_value.startswith("ark:/12148/"):
        short = ark_value.replace("ark:/12148/", "")
        return f"{GALLICA_BASE}/ark:/12148/{short}"
    if re.match(r"^[a-z0-9]+$", ark_value):
        return f"{GALLICA_BASE}/ark:/12148/{ark_value}"
    return ark_value


def get_issues_via_sru(ark_id: str, max_records: int = MAX_ISSUES) -> list[dict]:
    short_id  = ark_id.replace("ark:/12148/", "")
    ark_press = f"{short_id}_date"
    issues = []
    start_record = 1
    page_size = 50
    total = None

    while True:
        params = {
            "operation":      "searchRetrieve",
            "version":        "1.2",
            "query":          f'(dc.type all "fascicule") and arkPress all "{ark_press}"',
            "startRecord":    start_record,
            "maximumRecords": page_size,
            "collapsing":     "false",
        }
        try:
            resp = requests.get(GALLICA_SRU, params=params, headers=HEADERS, timeout=TIMEOUT)
            resp.raise_for_status()
        except requests.exceptions.Timeout:
            raise TimeoutError(f"Gallica SRU n'a pas répondu dans les {TIMEOUT}s.")
        except requests.exceptions.HTTPError as e:
            raise ConnectionError(f"Erreur HTTP {resp.status_code} : {e}")
        except requests.exceptions.RequestException as e:
            raise ConnectionError(f"Erreur réseau : {e}")

        xml = resp.text
        if start_record == 1:
            m = re.search(r"<numberOfRecords>(\d+)</numberOfRecords>", xml)
            total = int(m.group(1)) if m else 0
            if total == 0:
                break

        records = re.findall(r"<srw:record>(.*?)</srw:record>", xml, re.DOTALL)
        if not records:
            break

        for rec in records:
            def first(tag, text=rec):
                m = re.search(rf"<dc:{tag}[^>]*>(.*?)</dc:{tag}>", text, re.DOTALL)
                return m.group(1).strip() if m else ""
            raw_id = first("identifier")
            issues.append({
                "date":        first("date"),
                "titre":       first("title") or "(sans titre)",
                "description": first("description"),
                "url":         build_gallica_url(raw_id),
            })

        start_record += page_size
        if start_record > min(total, max_records):
            break
        time.sleep(0.3)

    return issues


def get_issues_via_issues_api(ark_id: str) -> list[dict]:
    base_url = f"{GALLICA_BASE}/services/Issues"
    try:
        r = requests.get(base_url, params={"ark": f"{ark_id}/date"}, headers=HEADERS, timeout=TIMEOUT)
        r.raise_for_status()
    except requests.exceptions.RequestException as e:
        raise ConnectionError(f"Erreur API Issues (années) : {e}")

    years = re.findall(r"<year>(\d{4})</year>", r.text)
    if not years:
        return []

    issues = []
    for year in years:
        try:
            r2 = requests.get(base_url, params={"ark": f"{ark_id}/date", "date": year}, headers=HEADERS, timeout=TIMEOUT)
            r2.raise_for_status()
        except requests.exceptions.RequestException:
            continue

        for m in re.finditer(r'<issue\b[^>]*\bark="([^"]+)"[^>]*>([^<]*)</issue>', r2.text):
            raw_ark = m.group(1).strip()
            label   = m.group(2).strip()
            issues.append({
                "date":        year,
                "titre":       label or f"Numéro de {year}",
                "description": "",
                "url":         build_gallica_url(raw_ark),
            })
        time.sleep(0.2)

    return issues


def make_xlsx(df: pd.DataFrame) -> bytes:
    """Génère un fichier xlsx formaté en mémoire et retourne les bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Numéros Gallica"

    # En-têtes
    headers = list(df.columns)
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h.capitalize())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    # Données
    row_font     = Font(name="Arial", size=10)
    url_font     = Font(name="Arial", size=10, color="1F4E79", underline="single")
    alt_fill     = PatternFill("solid", start_color="EBF3FB")

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = url_font if headers[col_idx - 1] == "url" else row_font
            if fill:
                cell.fill = fill

    # Largeurs de colonnes
    col_widths = {"date": 12, "titre": 45, "description": 40, "url": 55}
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 20)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Interface Streamlit ───────────────────────────────────────────────────────

st.set_page_config(page_title="Moissonneur Gallica", page_icon="📰", layout="wide")
st.title("📰 Moissonneur Gallica")
st.caption(
    "Entrez l'URL d'une revue sur [Gallica](https://gallica.bnf.fr) "
    "pour lister ses numéros disponibles."
)

url = st.text_input(
    "URL de la revue Gallica",
    placeholder="ex. : https://gallica.bnf.fr/ark:/12148/cb32731059c/date",
)

if url:
    ark = extract_ark(url)
    if not ark:
        st.error("❌ Impossible d'extraire un identifiant ARK. L'URL doit contenir `ark:/12148/…`.")
        st.stop()

    st.success(f"✅ ARK détecté : `{ark}`")

    issues = []
    method_used = ""

    with st.spinner("Interrogation de l'API Gallica SRU (méthode arkPress)…"):
        try:
            issues = get_issues_via_sru(ark)
            method_used = "SRU / arkPress"
        except Exception as e:
            st.warning(f"⚠️ SRU indisponible ({e}), bascule sur l'API Issues…")

    if not issues:
        with st.spinner("Interrogation de l'API Issues de Gallica…"):
            try:
                issues = get_issues_via_issues_api(ark)
                method_used = "API Issues (fallback)"
            except Exception as e:
                st.error(f"💥 Erreur : {e}")
                st.stop()

    if not issues:
        st.warning("⚠️ Aucun numéro trouvé.")
    else:
        df = pd.DataFrame(issues)[["date", "titre", "description", "url"]]

        # ── Boutons de téléchargement EN HAUT ──────────────────────────────
        st.metric("Numéros récupérés", len(issues))
        st.caption(f"Source : {method_used}")

        csv_data  = df.to_csv(index=False).encode("utf-8")
        xlsx_data = make_xlsx(df)

        col1, col2, _ = st.columns([1, 1, 4])
        with col1:
            st.download_button(
                label="⬇️ CSV",
                data=csv_data,
                file_name="gallica_issues.csv",
                mime="text/csv",
            )
        with col2:
            st.download_button(
                label="⬇️ Excel",
                data=xlsx_data,
                file_name="gallica_issues.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # ── Tableau ────────────────────────────────────────────────────────
        df_display = df.copy()
        df_display["url"] = df_display["url"].apply(
            lambda u: f'<a href="{u}" target="_blank">🔗 Voir</a>' if u else ""
        )
        st.write(df_display.to_html(escape=False, index=False), unsafe_allow_html=True)
